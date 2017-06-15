# -*- coding: utf-8 -*-
# pylint: disable=too-many-ancestors
"""Views for the shop application."""

import re
from subprocess import call

import openpyxl
from constance import config
from django.db.models import Count, F, Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets
from rest_framework.pagination import LimitOffsetPagination
from rest_framework.response import Response
from rest_framework.views import APIView

from . import serializers
from .models import Category, Item, Order, Purchase
from .tickets import print_ticket, print_total, update_printer


class ItemViewSet(viewsets.ModelViewSet):
    """Item viewset of the shop."""

    serializer_class = serializers.ItemSerializer
    queryset = Item.objects.all()


class PurchaseViewSet(viewsets.ModelViewSet):
    """Purchase viewset of the shop."""

    serializer_class = serializers.HistorySerializer
    queryset = Purchase.objects.all()


class StoreView(APIView):
    """List the items of the shop grouped by categories."""

    serializer_class = serializers.CategorySerializer

    def get(self, request):
        """GET request for items of the shop grouped by categories."""
        categories = [c for c in Category.objects.all() if c.item_set.all()]
        return Response(self.serializer_class(categories, many=True).data)


class HistoryView(APIView):
    """List the history of purchases of the shop."""

    serializer_class = serializers.HistorySerializer

    def get(self, request):
        """GET request for the list the history of purchases of the shop."""
        purchases = Purchase.objects.annotate(num_orders=Count('orders'))\
                                    .filter(num_orders__gt=0)
        paginator = LimitOffsetPagination()
        result_page = paginator.paginate_queryset(purchases, request)
        return Response(self.serializer_class(result_page, many=True).data)


class BuyView(APIView):
    """Buy items in the shop."""

    serializer_class = serializers.BuySerializer

    def post(self, request):
        """POST request to purcharse items in the shop."""
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        orders = serializer.validated_data['orders']
        receipt = serializer.validated_data['receipt']

        # ------------------------------------------------------------------- #
        # Verify that items are valid and haven't changed                     #
        # ------------------------------------------------------------------- #

        items = []
        for order in orders:
            try:
                item = Item.objects.get(id=order['item_id'])
            except Item.DoesNotExist:
                return Response({'status': 'Some items are not valid or '
                                           'have changed in the database.'},
                                status=400)
            items.append((item, order['quantity']))

        # ------------------------------------------------------------------- #
        # Print the tickets                                                   #
        # ------------------------------------------------------------------- #

        # Update the printer
        update_printer()

        for (item, quantity) in items:
            if item.individual_tickets and config.INDIVIDUAL_TICKETS:
                if item.group_tickets:
                    print_ticket(item, quantity)
                else:
                    for _iterate in range(quantity):
                        print_ticket(item)
        if receipt:
            print_total(items)
        if config.VENDOR_RECEIPT:
            print_total(items)

        # ------------------------------------------------------------------- #
        # Store the purchase                                                  #
        # ------------------------------------------------------------------- #

        purchase = Purchase(date=timezone.now())
        purchase.save()
        for (item, quantity) in items:
            Order(item=item,
                  purchase=purchase,
                  price=item.price,
                  quantity=quantity).save()

        return Response({'status': 'ok'})


class PrintReceiptView(APIView):
    """Print a receipt for a purchase."""

    serializer_class = serializers.PrintReceiptSerializer

    def post(self, request):
        """POST request to print a receipt."""
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        purchase_id = serializer.validated_data['purchase_id']

        try:
            purchase = Purchase.objects.get(pk=purchase_id)
        except Purchase.DoesNotExist:
            return Response({'status': 'failed'}, status=404)

        items = [(o.item, o.quantity)
                 for o in Order.objects.filter(purchase=purchase)]
        print_total(items)

        return Response({'status': 'ok'})


class StatsView(APIView):
    """The stats of the shop."""

    def get(self, request):
        """GET request to access the stats of the shop."""
        stats = dict()

        # ------------------------------------------------------------------- #
        # Total sales                                                         #
        # ------------------------------------------------------------------- #

        total_sales = Order.objects.aggregate(
            total=Sum(F('price') * F('quantity')))
        stats['total_sales'] = total_sales['total']

        # ------------------------------------------------------------------- #
        # Cumulative Sales                                                    #
        # ------------------------------------------------------------------- #

        purchases = Purchase.objects.order_by('date').annotate(
            total=Sum(F('orders__price')*F('orders__quantity')))
        running_total = 0
        stats['cumulative_sales'] = []
        for purchase in purchases:
            if purchase.total is None:
                purchase.delete()
                continue
            running_total += purchase.total
            stats['cumulative_sales'].append({
                'x': int(purchase.date.timestamp()),
                'y': running_total
            })

        # ------------------------------------------------------------------- #
        # Count                                                               #
        # ------------------------------------------------------------------- #

        items = Item.objects.annotate(count=Sum('orders__quantity'))
        stats['counts'] = dict()
        stats['counts']['labels'] = [i.name for i in items]
        stats['counts']['series'] = [i.count for i in items]

        return Response(stats)


class ShutdownView(APIView):
    """Shutdown the shop."""

    def get(self, request):
        """GET request to shutdown the shop."""
        call(["sudo", "systemctl", "poweroff"])
        return Response({'status': 'success'})


class ConfigView(APIView):
    """The config of the shop."""

    def get(self, request):
        """GET request to access the configuration of the shop."""
        keys = list(config.__dir__())
        confs = {k: config.__getattr__(k) for k in keys}
        return Response(confs)


class ExportView(APIView):
    """Export the shop."""

    def get(self, request):
        """GET request to export the shop."""
        # Prepare the data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = config.EVENT_NAME
        ws.row_dimensions[1].height = 22
        ws['A1'] = config.EVENT_NAME
        ws['A1'].font = openpyxl.styles.Font(size=18)
        ws['A1'].alignment = openpyxl.styles.Alignment(horizontal="center",
                                                       vertical="center")
        ws.merge_cells("A1:G1")
        ws.append(["Date", "Time", "Purchase", "Article",
                   "Article price", "Quantity", "Total price"])
        for row in ws.iter_rows("A2:G2"):
            for cell in row:
                cell.font = openpyxl.styles.Font(bold=True)
        for order in Order.objects.all():
            ws.append(["{:%Y-%m-%d}".format(order.purchase.date),
                       "{:%H:%M}".format(order.purchase.date),
                       order.purchase.id,
                       order.item.name,
                       order.quantity,
                       order.price,
                       order.price * order.quantity])
        for column_cells in ws.columns:
            ws.column_dimensions[column_cells[0].column].width = 15
        # Prepare the filename
        filename = "{:%Y-%m-%d_%H-%M}_{}.xlsx".format(
            timezone.localtime(timezone.now()),
            re.sub(r'\W+', '', config.EVENT_NAME))
        # Prepare the response as downloadable content
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = ('attachment;'
                                           'filename="{}"'.format(filename))
        wb.save(response)
        return response
